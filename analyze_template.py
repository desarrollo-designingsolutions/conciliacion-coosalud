import openpyxl
from pathlib import Path

wb = openpyxl.load_workbook('XXXXXXX_GUIA.xlsx')
ws = wb.active

print('=== INFORMACIÓN DEL ARCHIVO ===')
print(f'Hoja activa: {ws.title}')
print(f'Dimensiones: {ws.dimensions}')

print('\n=== IMÁGENES ENCONTRADAS ===')
if hasattr(ws, '_images') and ws._images:
    print(f'Total de imágenes: {len(ws._images)}')
    for idx, img in enumerate(ws._images):
        print(f'\nImagen {idx + 1}:')
        print(f'  Tipo: {type(img).__name__}')
        try:
            if hasattr(img, 'anchor'):
                if hasattr(img.anchor, '_from'):
                    print(f'  Desde: Fila {img.anchor._from.row}, Col {img.anchor._from.col}')
                if hasattr(img.anchor, '_to') and img.anchor._to:
                    print(f'  Hasta: Fila {img.anchor._to.row}, Col {img.anchor._to.col}')
                if hasattr(img.anchor, 'anchorType'):
                    print(f'  Tipo de ancla: {img.anchor.anchorType}')
        except Exception as e:
            print(f'  Error al leer anchor: {e}')
        
        if hasattr(img, 'path'):
            print(f'  Path: {img.path}')
        if hasattr(img, 'ref'):
            print(f'  Ref: {img.ref}')
else:
    print('No se encontraron imágenes embebidas')

print('\n=== BUSCANDO FÓRMULAS CON IMAGE ===')
formulas_found = False
for row in ws.iter_rows(min_row=1, max_row=35):
    for cell in row:
        if cell.value:
            cell_str = str(cell.value)
            if cell_str.startswith('='):
                if 'IMAGE' in cell_str.upper() or 'IMAGEN' in cell_str.upper():
                    print(f'  {cell.coordinate}: {cell.value}')
                    formulas_found = True

if not formulas_found:
    print('No se encontraron fórmulas IMAGE')

print('\n=== TODAS LAS CELDAS CON VALORES (Filas 1-35) ===')
for row_num in range(1, 36):
    for col_num in range(1, 14):
        cell = ws.cell(row=row_num, column=col_num)
        if cell.value:
            value_preview = str(cell.value)[:50]
            print(f'  {cell.coordinate}: {value_preview}')
