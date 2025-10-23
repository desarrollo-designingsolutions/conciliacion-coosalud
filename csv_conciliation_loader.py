#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# CSV -> Validaciones -> (opcional) MySQL insert masivo -> Reportes de errores / éxito
#
# Uso:
#   python csv_conciliation_loader.py --csv "/ruta/al/archivo.csv" --out-dir "/ruta/de/salida" [--verbose]
#
# Variables de entorno (para fase MySQL):
#   MYSQL_HOST, MYSQL_PORT (3306), MYSQL_DB, MYSQL_USER, MYSQL_PASSWORD

import argparse
import csv
import os
import sys
import logging
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path
from copy import copy
import uuid

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

try:
    import pymysql
except Exception:
    pymysql = None  # modo "solo validaciones" si no está instalado

# ------------------------ Config básica ------------------------

EXPECTED_HEADERS = [
    "ID",
    "FACTURA_ID",
    "SERVICIO_ID",
    "ORIGIN",
    "NIT",
    "RAZON_SOCIAL",
    "NUMERO_FACTURA",
    "FECHA_INICIO",
    "FECHA_FIN",
    "MODALIDAD",
    "REGIMEN",
    "COBERTURA",
    "CONTRATO",
    "TIPO_DOCUMENTO",
    "NUMERO_DOCUMENTO",
    "PRIMER_NOMBRE",
    "SEGUNDO_NOMBRE",
    "PRIMER_APELLIDO",
    "SEGUNDO_APELLIDO",
    "GENERO",
    "CODIGO_SERVICIO",
    "DESCRIPCION_SERVICIO",
    "CANTIDAD_SERVICIO",
    "VALOR_UNITARIO_SERVICIO",
    "VALOR_TOTAL_SERVICIO",
    "CODIGOS_GLOSA",
    "OBSERVACIONES_GLOSAS",
    "VALOR_GLOSA",
    "VALOR_APROBADO",
    "ESTADO_RESPUESTA",
    "NUMERO_DE_AUTORIZACION",
    "RESPUESTA_DE_IPS",
    "VALOR_ACEPTADO_POR_IPS",
    "VALOR_ACEPTADO_POR_EPS",
    "VALOR_RATIFICADO_EPS",
    "OBSERVACIONES",
]

TZ = ZoneInfo("America/Bogota")
ACTA_TEMPLATE_NAME = "XXXXXXX_GUIA.xlsx"
ACTA_OUTPUT_DIRNAME = "actas_conciliacion"

# códigos de salida
EXIT_OK = 0
EXIT_CSV_ERRORS = 1
EXIT_CSV_READ_ERROR = 2
EXIT_MYSQL_ERROR = 3

# ------------------------ Logging helpers ------------------------

def setup_logger(verbose: bool, log_file: str | None):
    level = logging.DEBUG if verbose else logging.INFO
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    datefmt = "%Y-%m-%d %H:%M:%S"
    handlers = [logging.StreamHandler(sys.stdout)]
    if log_file:
        try:
            os.makedirs(os.path.dirname(log_file), exist_ok=True)
            handlers.append(logging.FileHandler(log_file, encoding="utf-8"))
        except Exception:
            pass
    logging.basicConfig(level=level, format=fmt, datefmt=datefmt, handlers=handlers)

def info(msg: str): logging.info(msg)
def debug(msg: str): logging.debug(msg)
def error(msg: str): logging.error(msg)

# ------------------------ Modelos ------------------------

@dataclass
class ErrorRow:
    consecutivo: int
    id_: str
    factura_id: str
    origin: str
    nit: str
    razon_social: str
    numero_factura: str
    fila: int
    columna: str
    valor: str
    error: str
    validacion: str

# ------------------------ Utilidades ------------------------

def now_bogota_str() -> str:
    dt = datetime.now(TZ)
    # dia-mes-año hora:segundos
    return dt.strftime("%d-%m-%Y %H:%S")

def is_uuid_hyphenated(s: str) -> bool:
    """
    Valida que s sea un UUID canónico con guiones (36 chars):
    xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx
    """
    if s is None:
        return False
    s = str(s).strip()
    try:
        u = uuid.UUID(s)
        return str(u) == s.lower()
    except Exception:
        return False

def parse_decimal_maybe_comma(x: str) -> Decimal:
    """
    Convierte '1.234,56' | '1234,56' | '1234.56' a Decimal.
    Si contiene ',' se asume coma como decimal (se quitan puntos de miles).
    """
    if x is None:
        raise InvalidOperation("valor nulo")
    s = str(x).strip()
    if s == "":
        raise InvalidOperation("vacío")
    s = s.replace(" ", "").replace("$", "")
    if "," in s:
        s2 = s.replace(".", "")
        s2 = s2.replace(",", ".")
        return Decimal(s2)
    else:
        return Decimal(s)

def write_errors_csv(errors: list[ErrorRow], out_path: str) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "Consecutivo","Id","Factura_id","Origin","Nit","Razón_social",
            "Numero_factura","Fila","Columna","Valor","Error","Validación"
        ])
        for e in errors:
            w.writerow([
                e.consecutivo, e.id_, e.factura_id, e.origin, e.nit, e.razon_social,
                e.numero_factura, e.fila, e.columna, e.valor, e.error, e.validacion
            ])


def generate_acta_excels(df: pd.DataFrame, ids: list[str], out_dir: str, conn) -> list[dict]:
    """Genera archivos de acta a partir de la plantilla y retorna metadatos de descarga."""
    if not ids or df.empty:
        return []

    df_subset = df[df["ID"].isin(ids)].copy()
    if df_subset.empty:
        return []

    template_path = Path(__file__).resolve().parent / ACTA_TEMPLATE_NAME
    if not template_path.exists():
        error(f"Plantilla de acta no encontrada: {template_path}")
        return []

    currency_format = '[$-es-CO]"$" #.##0'
    fallback_currency_format = '[$-en-US]"$" #,##0'

    acta_dir = Path(out_dir) / ACTA_OUTPUT_DIRNAME
    try:
        acta_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        error(f"No se pudo crear el directorio para actas ({acta_dir}): {exc}")
        return []

    generated_records: list[dict] = []
    usuario_default = os.getenv("ACTA_GENERATION_USER") or os.getenv("APP_USER") or os.getenv("USER") or "system"

    current_dt = datetime.now(TZ)
    current_dt_naive = current_dt.replace(tzinfo=None)
    month_names = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
        7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
    }

    df_subset["NIT"] = df_subset["NIT"].astype(str).str.strip()
    df_subset["RAZON_SOCIAL"] = df_subset["RAZON_SOCIAL"].astype(str).str.strip()

    format_error_reported = False
    currency_cells: list = []

    def apply_currency(cell):
        nonlocal format_error_reported
        currency_cells.append(cell)
        if format_error_reported:
            cell.number_format = fallback_currency_format
            return
        try:
            cell.number_format = currency_format
        except Exception as exc:
            if not format_error_reported:
                error(f"No se pudo aplicar formato monetario personalizado: {exc}")
                format_error_reported = True
            cell.number_format = fallback_currency_format

    location_cache: dict[str, tuple[str, str]] = {}
    column_cache: dict[str, set[str]] = {}

    def get_db_name() -> str:
        if conn is None:
            return ""
        db_name = getattr(conn, "db", "") or ""
        if isinstance(db_name, bytes):
            db_name = db_name.decode("utf-8", errors="ignore")
        if not db_name:
            db_name = os.getenv("MYSQL_DB", "")
        return db_name

    def get_table_columns(table: str) -> set[str]:
        if conn is None:
            return set()
        table_key = table.lower()
        if table_key in column_cache:
            return column_cache[table_key]
        db_name = get_db_name()
        if not db_name:
            column_cache[table_key] = set()
            return set()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT COLUMN_NAME
                    FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                    """,
                    (db_name, table),
                )
                cols = {row[0].lower() for row in cur.fetchall()}
                column_cache[table_key] = cols
                return cols
        except Exception as exc:
            error(f"No se pudieron obtener las columnas de {table}: {exc}")
            column_cache[table_key] = set()
            return set()

    def fetch_location(nit: str) -> tuple[str, str]:
        key = nit or ""
        if key in location_cache:
            return location_cache[key]
        departamento = ""
        municipio = ""
        if conn is not None and key:
            try:
                with conn.cursor() as cur:
                    row = None
                    td_columns = get_table_columns("third_departments")
                    if "nit" in td_columns:
                        cur.execute(
                            """
                            SELECT departamento, municipio
                            FROM third_departments
                            WHERE nit = %s
                            ORDER BY updated_at DESC, id DESC
                            LIMIT 1
                            """,
                            (key,),
                        )
                        row = cur.fetchone()

                    if row is None and "third_id" in td_columns:
                        third_columns = get_table_columns("thirds")
                        conditions: list[str] = []
                        params: list[str] = []
                        candidate_map = [
                            ("nit", "t.nit = %s"),
                            ("identification", "t.identification = %s"),
                            ("numero_identificacion", "t.numero_identificacion = %s"),
                            ("documento", "t.documento = %s"),
                            ("document_number", "t.document_number = %s"),
                        ]
                        for col_name, expr in candidate_map:
                            if col_name in third_columns:
                                conditions.append(expr)
                                params.append(key)

                        if conditions:
                            where_clause = " OR ".join(conditions)
                            query = f"""
SELECT td.departamento, td.municipio
FROM third_departments td
INNER JOIN thirds t ON t.id = td.third_id
WHERE {where_clause}
ORDER BY td.updated_at DESC, td.id DESC
LIMIT 1
"""
                            cur.execute(query, tuple(params))
                            row = cur.fetchone()

                    if row:
                        departamento = (row[0] or "").strip()
                        municipio = (row[1] or "").strip()
            except Exception as exc:
                error(f"No se pudo obtener departamento/municipio para el NIT {key}: {exc}")
        location_cache[key] = (departamento, municipio)
        return location_cache[key]

    def parse_decimal(value: str | Decimal | None) -> int:
        if value in (None, ""):
            return 0
        try:
            dec = parse_decimal_maybe_comma(value)
            dec = dec.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            return int(dec)
        except Exception:
            return 0

    def parse_fecha(value: str | None):
        if not value:
            return None
        text = str(value).strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    unique_pairs = df_subset[["NIT", "RAZON_SOCIAL"]].drop_duplicates()

    for _, pair in unique_pairs.iterrows():
        nit = pair.get("NIT", "")
        razon_social = pair.get("RAZON_SOCIAL", "")
        pair_rows = df_subset[(df_subset["NIT"] == nit) & (df_subset["RAZON_SOCIAL"] == razon_social)].copy()
        if pair_rows.empty:
            continue

        currency_cells.clear()
        departamento, municipio = fetch_location(nit)
        total_glosa = Decimal('0')
        total_eps = Decimal('0')
        total_ips = Decimal('0')
        total_rat = Decimal('0')

        wb = None
        try:
            wb = load_workbook(template_path)
            ws = wb.active

            for col_idx in range(1, 14):
                header_cell = ws.cell(row=7, column=col_idx)
                try:
                    header_cell.font = header_cell.font.copy(name="Calibri", size=5)
                except Exception:
                    header_cell.font = Font(name="Calibri", size=5)

            ws["F5"] = nit
            ws["H4"] = razon_social
            ws["J5"] = current_dt_naive
            ws["J5"].number_format = "DD/MM/YYYY"
            ws["B4"] = departamento
            ws["B5"] = municipio

            row_count = len(pair_rows)
            if row_count == 0:
                continue

            additional_rows = max(row_count - 1, 0)
            if additional_rows:
                ws.insert_rows(9, amount=additional_rows)

            template_cells = tuple(ws[8])
            for offset in range(row_count):
                target_row = 8 + offset
                for src_cell in template_cells:
                    dest_cell = ws.cell(row=target_row, column=src_cell.col_idx)
                    if src_cell.has_style:
                        try:
                            dest_cell._style = copy(src_cell._style)
                        except Exception:
                            dest_cell.font = src_cell.font
                            dest_cell.border = src_cell.border
                            dest_cell.fill = src_cell.fill
                            dest_cell.number_format = src_cell.number_format
                            dest_cell.protection = src_cell.protection
                            dest_cell.alignment = src_cell.alignment
                    dest_cell.value = None

            data_start_row = 8
            pair_rows = pair_rows.reset_index(drop=True)
            for idx_row, record in pair_rows.iterrows():
                excel_row = data_start_row + idx_row
                numero_factura = str(record.get("NUMERO_FACTURA", "")).strip()
                codigo_glosa = str(record.get("CODIGOS_GLOSA", "")).strip()
                contrato = str(record.get("CONTRATO", "")).strip()
                valor_factura = parse_decimal(record.get("VALOR_TOTAL_SERVICIO"))
                fecha_inicio = record.get("FECHA_INICIO", "")
                fecha_dt = parse_fecha(fecha_inicio)
                valor_glosa = parse_decimal(record.get("VALOR_GLOSA"))
                try:
                    total_glosa += parse_decimal_maybe_comma(str(record.get("VALOR_GLOSA", "0")))
                except Exception:
                    pass
                valor_eps = parse_decimal(record.get("VALOR_ACEPTADO_POR_EPS"))
                try:
                    total_eps += parse_decimal_maybe_comma(str(record.get("VALOR_ACEPTADO_POR_EPS", "0")))
                except Exception:
                    pass
                valor_ips = parse_decimal(record.get("VALOR_ACEPTADO_POR_IPS"))
                try:
                    total_ips += parse_decimal_maybe_comma(str(record.get("VALOR_ACEPTADO_POR_IPS", "0")))
                except Exception:
                    pass
                valor_rat = parse_decimal(record.get("VALOR_RATIFICADO_EPS"))
                try:
                    total_rat += parse_decimal_maybe_comma(str(record.get("VALOR_RATIFICADO_EPS", "0")))
                except Exception:
                    pass
                justificacion = str(record.get("OBSERVACIONES", "")).strip()

                ws.cell(row=excel_row, column=1, value=numero_factura)
                ws.cell(row=excel_row, column=2, value=numero_factura)
                ws.cell(row=excel_row, column=3, value=codigo_glosa)
                ws.cell(row=excel_row, column=4, value=contrato)
                cell_valor_factura = ws.cell(row=excel_row, column=5, value=valor_factura)
                apply_currency(cell_valor_factura)

                if fecha_dt:
                    cell_fecha = ws.cell(row=excel_row, column=6, value=fecha_dt)
                    cell_fecha.number_format = "DD/MM/YYYY"
                else:
                    ws.cell(row=excel_row, column=6, value=str(fecha_inicio))

                ws.cell(row=excel_row, column=7, value=departamento)
                cell_valor_glosa = ws.cell(row=excel_row, column=8, value=valor_glosa)
                apply_currency(cell_valor_glosa)
                cell_pendiente = ws.cell(row=excel_row, column=9, value=0)
                apply_currency(cell_pendiente)
                cell_valor_eps = ws.cell(row=excel_row, column=10, value=valor_eps)
                apply_currency(cell_valor_eps)
                cell_valor_ips = ws.cell(row=excel_row, column=11, value=valor_ips)
                apply_currency(cell_valor_ips)
                cell_valor_rat = ws.cell(row=excel_row, column=12, value=valor_rat)
                apply_currency(cell_valor_rat)
                ws.cell(row=excel_row, column=13, value=justificacion)

            data_end_row = data_start_row + row_count - 1
            row_shift = row_count - 1
            results_row = 9 + row_shift

            for col in ("E", "H", "I", "J", "K", "L"):
                cell = ws[f"{col}{results_row}"]
                cell.value = f"=SUM({col}{data_start_row}:{col}{data_end_row})"
                apply_currency(cell)

            for offset, source_col in enumerate(["E", "H", "I", "J", "K", "L"], start=12):
                target_cell = ws[f"C{offset + row_shift}"]
                target_cell.value = f"={source_col}{results_row}"
                apply_currency(target_cell)

            footer_row = 19 + row_shift
            month_name = month_names.get(current_dt.month, "")
            footer_text = (
                "La presente acta se expide en la ciudad de CARTAGENA, el día "
                f"{current_dt.day} del mes de {month_name} de {current_dt.year} "
                "y se suscribe por los funcionarios representates de las entidades que participan en el proceso de conciliación."
            )
            ws[f"A{footer_row}"] = footer_text

            timestamp = current_dt.strftime("%Y%m%d_%H%M%S")
            safe_nit = ''.join(ch for ch in str(nit) if ch.isalnum()) or "sin_nit"
            out_path = acta_dir / f"acta_conciliacion_{safe_nit}_{timestamp}.xlsx"
            try:
                wb.save(out_path)
            except TypeError as exc:
                if not format_error_reported:
                    error(
                        "No se pudo guardar el acta con el formato monetario personalizado; se usará el formato alterno. "
                        f"Detalle: {exc}"
                    )
                    format_error_reported = True
                for cell in currency_cells:
                    try:
                        cell.number_format = fallback_currency_format
                    except Exception:
                        pass
                wb.save(out_path)

            sum_glosa = total_glosa.quantize(Decimal('0.01'))
            sum_eps = total_eps.quantize(Decimal('0.01'))
            sum_ips = total_ips.quantize(Decimal('0.01'))
            sum_rat = total_rat.quantize(Decimal('0.01'))

            record_info = {
                'nit': nit,
                'razon_social': razon_social,
                'file_name': out_path.name,
                'file_path': str(out_path),
                'download_link': str(out_path),
                'valor_glosa': str(sum_glosa),
                'valor_aceptado_eps': str(sum_eps),
                'valor_aceptado_ips': str(sum_ips),
                'valor_ratificado': str(sum_rat),
            }
            generated_records.append(record_info)

            if conn is not None:
                try:
                    import uuid
                    acta_id = str(uuid.uuid4())
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                                INSERT INTO conciliation_acta_files
                                (id, nit, razon_social, file_name, file_path, valor_glosa, valor_aceptado_eps,
                                 valor_aceptado_ips, valor_ratificado, usuario, created_at, updated_at)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, NOW(), NOW())
                            """,
                            (
                                acta_id,
                                nit,
                                razon_social,
                                out_path.name,
                                str(out_path),
                                str(sum_glosa),
                                str(sum_eps),
                                str(sum_ips),
                                str(sum_rat),
                                usuario_default,
                            ),
                        )
                    conn.commit()
                except Exception as db_exc:
                    error(f"No se pudo registrar el acta en conciliation_acta_files: {db_exc}")
        except Exception as exc:
            error(f"No se pudo generar acta para el NIT {nit or 'desconocido'}: {exc}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    return generated_records

def get_mysql_conn_from_env():
    host = os.getenv("MYSQL_HOST")
    port = int(os.getenv("MYSQL_PORT", "3306"))
    db = os.getenv("MYSQL_DB")
    user = os.getenv("MYSQL_USER")
    pw = os.getenv("MYSQL_PASSWORD")
    if not all([host, db, user, pw]):
        return None, "Variables de entorno MySQL incompletas (MYSQL_HOST, MYSQL_DB, MYSQL_USER, MYSQL_PASSWORD)"
    if pymysql is None:
        return None, "pymysql no instalado. Instala con: pip install pymysql"
    try:
        info(f"Conectando a MySQL en {host}:{port}/{db} …")
        conn = pymysql.connect(host=host, port=port, user=user, password=pw,
                               database=db, autocommit=False, charset="utf8mb4")
        info("Conexión MySQL OK.")
        # Forzar colación de la sesión para evitar "Illegal mix of collations"
        with conn.cursor() as cur:
            cur.execute("SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci")
            cur.execute("SET collation_connection = 'utf8mb4_unicode_ci'")
        conn.commit()

        return conn, None
    except Exception as ex:
        return None, f"No fue posible conectar a MySQL: {ex}"


        

# ------------------------ Validaciones CSV ------------------------

def validate_csv(df: pd.DataFrame, headers_in_file: list[str]) -> list[ErrorRow]:
    errors: list[ErrorRow] = []
    consecutivo = 1

    # 1) Encabezados exactos y en el mismo orden
    if headers_in_file != EXPECTED_HEADERS:
        now_s = now_bogota_str()
        errors.append(ErrorRow(
            consecutivo, "", "", "", "", "", "",
            1, "HEADERS", "|".join(headers_in_file),
            "Encabezados inválidos o fuera de orden. Deben coincidir exactamente con la especificación.",
            now_s
        ))
        return errors  # no seguimos si headers están mal

    # 2) ID no vacío, UUID con guiones canónico y sin duplicados
    seen_ids = set()
    for idx, row in df.iterrows():
        fila_excel = idx + 2  # encabezado en fila 1
        now_s = now_bogota_str()
        id_val = str(row.get("ID", "")).strip()
        if id_val == "":
            errors.append(ErrorRow(consecutivo, id_val, str(row.get("FACTURA_ID","")), str(row.get("ORIGIN","")),
                                   str(row.get("NIT","")), str(row.get("RAZON_SOCIAL","")), str(row.get("NUMERO_FACTURA","")),
                                   fila_excel, "ID", id_val, "ID vacío.", now_s)); consecutivo += 1
        elif not is_uuid_hyphenated(id_val):
            errors.append(ErrorRow(consecutivo, id_val, str(row.get("FACTURA_ID","")), str(row.get("ORIGIN","")),
                                   str(row.get("NIT","")), str(row.get("RAZON_SOCIAL","")), str(row.get("NUMERO_FACTURA","")),
                                   fila_excel, "ID", id_val,
                                   "ID debe ser UUID válido con guiones (xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx).", now_s)); consecutivo += 1
        elif id_val in seen_ids:
            errors.append(ErrorRow(consecutivo, id_val, str(row.get("FACTURA_ID","")), str(row.get("ORIGIN","")),
                                   str(row.get("NIT","")), str(row.get("RAZON_SOCIAL","")), str(row.get("NUMERO_FACTURA","")),
                                   fila_excel, "ID", id_val, "ID duplicado en el CSV.", now_s)); consecutivo += 1
        else:
            seen_ids.add(id_val)

    # 3) FACTURA_ID no vacío
    for idx, row in df.iterrows():
        fila_excel = idx + 2
        now_s = now_bogota_str()
        factura_id = str(row.get("FACTURA_ID", "")).strip()
        if factura_id == "":
            errors.append(ErrorRow(consecutivo, str(row.get("ID","")), factura_id, str(row.get("ORIGIN","")),
                                   str(row.get("NIT","")), str(row.get("RAZON_SOCIAL","")), str(row.get("NUMERO_FACTURA","")),
                                   fila_excel, "FACTURA_ID", factura_id, "FACTURA_ID vacío.", now_s)); consecutivo += 1

    # 4) Numéricos y suma (IPS + EPS + RATIFICADO == VALOR_GLOSA)
    for idx, row in df.iterrows():
        fila_excel = idx + 2
        now_s = now_bogota_str()

        def report(col, val, msg):
            nonlocal consecutivo
            errors.append(ErrorRow(consecutivo, str(row.get("ID","")), str(row.get("FACTURA_ID","")), str(row.get("ORIGIN","")),
                                   str(row.get("NIT","")), str(row.get("RAZON_SOCIAL","")), str(row.get("NUMERO_FACTURA","")),
                                   fila_excel, col, str(val), msg, now_s))
            consecutivo += 1

        cols_req = ["VALOR_ACEPTADO_POR_IPS", "VALOR_ACEPTADO_POR_EPS", "VALOR_RATIFICADO_EPS", "VALOR_GLOSA"]
        parsed: dict[str, Decimal | None] = {}
        for col in cols_req:
            val = row.get(col, "")
            try:
                parsed[col] = parse_decimal_maybe_comma(str(val))
            except InvalidOperation:
                report(col, val, f"{col} debe ser numérico (permite decimales con coma).")
                parsed[col] = None

        if all(parsed.get(c) is not None for c in cols_req):
            suma = parsed["VALOR_ACEPTADO_POR_IPS"] + parsed["VALOR_ACEPTADO_POR_EPS"] + parsed["VALOR_RATIFICADO_EPS"]
            if suma != parsed["VALOR_GLOSA"]:
                report("VALOR_GLOSA", row.get("VALOR_GLOSA",""),
                       "La suma de (IPS + EPS + RATIFICADO) debe ser igual a VALOR_GLOSA.")

    # 5) OBSERVACIONES no vacío
    for idx, row in df.iterrows():
        fila_excel = idx + 2
        now_s = now_bogota_str()
        obs = str(row.get("OBSERVACIONES", "")).strip()
        if obs == "":
            errors.append(ErrorRow(consecutivo, str(row.get("ID","")), str(row.get("FACTURA_ID","")), str(row.get("ORIGIN","")),
                                   str(row.get("NIT","")), str(row.get("RAZON_SOCIAL","")), str(row.get("NUMERO_FACTURA","")),
                                   fila_excel, "OBSERVACIONES", obs, "OBSERVACIONES debe estar diligenciado.", now_s)); consecutivo += 1

    return errors

# ------------------------ Errores SQL → CSV ------------------------

def _sql_error_rows_for_ids(df: pd.DataFrame, ids: list[str], msg: str) -> list[ErrorRow]:
    """
    Construye ErrorRow por cada ID afectado por un error SQL (conciliado previo o excepción).
    Usa la posición real de la fila para calcular 'Fila' (encabezado = 1).
    """
    errors: list[ErrorRow] = []
    consecutivo = 1
    now_s = now_bogota_str()

    for id_ in ids:
        subset = df.loc[df["ID"] == id_]
        if subset.empty:
            errors.append(ErrorRow(
                consecutivo, id_, "", "", "", "", "",
                0, "SQL", id_, msg, now_s
            ))
            consecutivo += 1
            continue

        pos = int(subset.index[0])
        fila_excel = pos + 2  # encabezado es fila 1
        row = subset.iloc[0]

        errors.append(ErrorRow(
            consecutivo,
            id_,
            str(row.get("FACTURA_ID", "")),
            str(row.get("ORIGIN", "")),
            str(row.get("NIT", "")),
            str(row.get("RAZON_SOCIAL", "")),
            str(row.get("NUMERO_FACTURA", "")),
            fila_excel,
            "SQL",
            id_,
            msg,
            now_s
        ))
        consecutivo += 1

    return errors

# ------------------------ Fase MySQL (masiva y rápida) ------------------------

def mysql_phase_and_success_report(df: pd.DataFrame, out_dir: str, ids: list[str]) -> tuple[bool, str, str]:
    """
    Inserta masivamente en conciliation_results usando tabla temporal y un único INSERT ... SELECT.
    Luego genera el CSV del JOIN de éxito. En caso de error, loguea por-ID en log_errores.csv.
    Retorna (ok, mensaje, path_del_csv_exito).
    """
    conn, err = get_mysql_conn_from_env()
    if conn is None:
        errors = _sql_error_rows_for_ids(df, ids, f"Fase MySQL omitida: {err}")
        out_err = os.path.join(out_dir, "log_errores.csv")
        write_errors_csv(errors, out_err)
        return False, f"Fase MySQL omitida: {err}", ""

    try:
        with conn.cursor() as cur:
            # 0) Filtrar IDs ya conciliados para no intentarlos de nuevo
            debug("Verificando conciliaciones existentes en lotes…")
            existing = set()
            chunk_size = 1000
            for i in range(0, len(ids), chunk_size):
                chunk = ids[i:i+chunk_size]
                placeholders = ",".join(["%s"] * len(chunk))
                cur.execute(
                    f"SELECT auditory_final_report_id FROM conciliation_results WHERE auditory_final_report_id IN ({placeholders})",
                    chunk
                )
                existing.update(row[0] for row in cur.fetchall())

            already_conciliated = sorted(list(existing))
            had_already_conciliated = bool(already_conciliated)
            ids_to_process = [i for i in ids if i not in existing]

            if already_conciliated:
                info(f"IDs ya conciliados detectados: {len(already_conciliated)}")
                # Registrar errores por cada uno
                msg = "Ese servicio ya se encuentra conciliado."
                errors = _sql_error_rows_for_ids(df, already_conciliated, msg)
                out_err = os.path.join(out_dir, "log_errores.csv")
                write_errors_csv(errors, out_err)
                if not ids_to_process:
                    info("Todos los servicios del CSV ya estaban conciliados; no se generarán actas.")
                    return False, "Las facturas cargadas ya se encuentran conciliadas en el sistema.", ""

            # 1) Preparar los registros a insertar desde el DataFrame (solo IDs a procesar)
            info(f"Preparando datos para inserción masiva (pendientes: {len(ids_to_process)})…")

            tmp_rows = []
            df_sub = df[df["ID"].isin(ids_to_process)].copy()

            for _idx, row in df_sub.iterrows():
                afr_id = str(row["ID"]).strip()
                resp_status = (str(row.get("ESTADO_RESPUESTA","")).strip() or None)
                autorization = (str(row.get("NUMERO_DE_AUTORIZACION","")).strip() or None)
                # Normalizar decimales
                try:
                    ips = parse_decimal_maybe_comma(row.get("VALOR_ACEPTADO_POR_IPS",""))
                    eps = parse_decimal_maybe_comma(row.get("VALOR_ACEPTADO_POR_EPS",""))
                    rat = parse_decimal_maybe_comma(row.get("VALOR_RATIFICADO_EPS",""))
                except InvalidOperation:
                    # Si algo raro pasó (no debería), saltar ese registro
                    continue
                ips_str = format(ips, 'f')
                eps_str = format(eps, 'f')
                rat_str = format(rat, 'f')
                observation = (str(row.get("OBSERVACIONES","")).strip() or None)

                tmp_rows.append((afr_id, resp_status, autorization, ips_str, eps_str, rat_str, observation))

            if not tmp_rows:
                return False, "No hay filas válidas para insertar (tras filtrar y normalizar).", ""

            # 2) Crear tabla temporal y cargar por lotes con colación fija
            cur.execute("""
                CREATE TEMPORARY TABLE IF NOT EXISTS tmp_conciliation_in (
                  afr_id CHAR(36) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL,
                  response_status VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NULL,
                  autorization_number VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NULL,
                  accepted_value_ips DECIMAL(18,2) NOT NULL,
                  accepted_value_eps DECIMAL(18,2) NOT NULL,
                  eps_ratified_value DECIMAL(18,2) NOT NULL,
                  observation TEXT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NULL,
                  PRIMARY KEY (afr_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            cur.execute("TRUNCATE TABLE tmp_conciliation_in")

            info("Cargando datos en tabla temporal…")
            insert_tmp_sql = """
                INSERT INTO tmp_conciliation_in
                (afr_id, response_status, autorization_number, accepted_value_ips, accepted_value_eps, eps_ratified_value, observation)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            batch_size = 5000
            total = len(tmp_rows)
            for i in range(0, total, batch_size):
                chunk = tmp_rows[i:i+batch_size]
                cur.executemany(insert_tmp_sql, chunk)
                info(f"Cargadas {min(i+batch_size, total)}/{total} filas en tmp_conciliation_in…")
            conn.commit()
            debug("Commit de carga temporal completado.")

            # 3) ÚNICO INSERT … SELECT (rápido) con colaciones alineadas
            info("Ejecutando inserción masiva en conciliation_results…")
            insert_bulk_sql = """
INSERT INTO conciliation_results
(
  id,
  auditory_final_report_id,
  invoice_audit_id,
  reconciliation_group_id,
  response_status,
  autorization_number,
  accepted_value_ips,
  accepted_value_eps,
  eps_ratified_value,
  created_at,
  updated_at,
  observation
)
SELECT
  UUID() AS id,
  afr.id AS auditory_final_report_id,
  afr.factura_id AS invoice_audit_id,
  (
    SELECT rg2.id
    FROM reconciliation_groups rg2
    WHERE rg2.third_id COLLATE utf8mb4_unicode_ci = afr.nit COLLATE utf8mb4_unicode_ci
    ORDER BY rg2.created_at DESC
    LIMIT 1
  ) AS reconciliation_group_id,
  tmp.response_status,
  tmp.autorization_number,
  tmp.accepted_value_ips,
  tmp.accepted_value_eps,
  tmp.eps_ratified_value,
  NOW() AS created_at,
  NOW() AS updated_at,
  tmp.observation
FROM tmp_conciliation_in tmp
INNER JOIN auditory_final_reports afr
  ON afr.id COLLATE utf8mb4_unicode_ci = tmp.afr_id COLLATE utf8mb4_unicode_ci
INNER JOIN invoice_audits ia
  ON ia.id COLLATE utf8mb4_unicode_ci = afr.factura_id COLLATE utf8mb4_unicode_ci
LEFT JOIN conciliation_results cr
  ON cr.auditory_final_report_id COLLATE utf8mb4_unicode_ci = afr.id COLLATE utf8mb4_unicode_ci
WHERE cr.id IS NULL
"""
            cur.execute(insert_bulk_sql)
            inserted = cur.rowcount  # filas insertadas
            info(f"Inserción masiva completada. Registros insertados: {inserted}")

            if inserted > 0:
                info("Actualizando tabla thirds_summary_conciliation con los nuevos valores…")
                update_summary_sql = """
INSERT INTO thirds_summary_conciliation (
  nit,
  valor_aceptado_eps,
  valor_aceptado_ips,
  valor_ratificado
)
SELECT
  ia.third_id AS nit,
  SUM(tmp.accepted_value_eps) AS valor_aceptado_eps,
  SUM(tmp.accepted_value_ips) AS valor_aceptado_ips,
  SUM(tmp.eps_ratified_value) AS valor_ratificado
FROM tmp_conciliation_in tmp
INNER JOIN auditory_final_reports afr
  ON afr.id COLLATE utf8mb4_unicode_ci = tmp.afr_id COLLATE utf8mb4_unicode_ci
INNER JOIN invoice_audits ia
  ON ia.id COLLATE utf8mb4_unicode_ci = afr.factura_id COLLATE utf8mb4_unicode_ci
GROUP BY ia.third_id
ON DUPLICATE KEY UPDATE
  valor_aceptado_eps = valor_aceptado_eps + VALUES(valor_aceptado_eps),
  valor_aceptado_ips = valor_aceptado_ips + VALUES(valor_aceptado_ips),
  valor_ratificado  = valor_ratificado  + VALUES(valor_ratificado)
"""
                cur.execute(update_summary_sql)
                info("Resumen actualizado correctamente.")

            conn.commit()

            # 4) Generar CSV de éxito (JOIN filtrado por IDs del CSV)
            # Si hubo IDs ya conciliados, omitir generación de actas.
            if 'had_already_conciliated' in locals() and had_already_conciliated:
                info("Se omite la generación de actas por detectar IDs ya conciliados en la carga.")
                return _generate_success_join(conn, out_dir, ids)
            else:
                return _finalize_success(conn, df, out_dir, ids)

    except Exception as ex:
        # Registrar el error SQL por cada ID que intentábamos procesar
        err_msg = f"Error durante fase MySQL: {ex}"
        errors = _sql_error_rows_for_ids(df, ids, err_msg)
        out_err = os.path.join(out_dir, "log_errores.csv")
        write_errors_csv(errors, out_err)
        try:
            conn.rollback()
        except Exception:
            pass
        return False, err_msg, ""
    finally:
        try:
            conn.close()
        except Exception:
            pass

# ------------------------ Reporte de éxito ------------------------

def _finalize_success(conn, df: pd.DataFrame, out_dir: str, ids: list[str]) -> tuple[bool, str, str]:
    try:
        ok, message, success_path = _generate_success_join(conn, out_dir, ids)
    except Exception as exc:
        return False, f"Error generando reporte de éxito: {exc}", ''
    if ok:
        try:
            acta_records = generate_acta_excels(df, ids, out_dir, conn)
            if acta_records:
                info("Actas de conciliación generadas:")
                for record in acta_records:
                    info(f"  - NIT {record.get('nit')}: {record.get('download_link')}")
        except Exception as exc:
            error(f"No se pudieron generar las actas de conciliación: {exc}")
    return ok, message, success_path

def _generate_success_join(conn, out_dir: str, ids: list[str]) -> tuple[bool, str, str]:
    """
    Genera CSV con JOIN de conciliation_results + auditory_final_reports + invoice_audits,
    filtrado por los IDs del CSV. Usa columnas reales de auditory_final_reports.
    """
    try:
        with conn.cursor() as cur:
            placeholders = ",".join(["%s"] * len(ids))
            join_sql = f"""
SELECT
  afr.id AS ID,
  afr.factura_id AS FACTURA_ID,
  afr.servicio_id AS SERVICIO_ID,
  afr.nit AS NIT,
  afr.razon_social AS RAZON_SOCIAL,
  afr.numero_factura AS NUMERO_FACTURA,
  afr.codigos_glosa AS CODIGOS_GLOSA,
  afr.valor_glosa AS VALOR_GLOSA,
  cs.id AS CONCILIATION_ID,
  cs.accepted_value_eps AS ACCEPTED_VALUE_EPS,
  cs.accepted_value_ips AS ACCEPTED_VALUE_IPS,
  cs.eps_ratified_value AS EPS_RATIFIED_VALUE
FROM conciliation_results cs
INNER JOIN auditory_final_reports afr
  ON afr.id = cs.auditory_final_report_id
INNER JOIN invoice_audits ia
  ON ia.id = afr.factura_id
WHERE afr.id IN ({placeholders})
ORDER BY afr.id
"""
            cur.execute(join_sql, ids)
            rows = cur.fetchall()
            cols = [desc[0] for desc in cur.description]

        success_path = os.path.join(out_dir, "carga_exitosa_join.csv")
        with open(success_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(cols)
            w.writerows(rows)

        return True, f"Proceso finalizado. Registros en reporte: {len(rows)}.", success_path
    except Exception as ex:
        return False, f"Error generando reporte de éxito: {ex}", ""
# ------------------------ main ------------------------

def _write_dry_run_summary(df: pd.DataFrame, out_dir: str, ids_to_insert: list[str]) -> str:
    """
    Genera un resumen por NIT/RAZON_SOCIAL con columnas solicitadas y lo guarda
    en out_dir/dry_run_summary.csv. Retorna la ruta del archivo.
    """
    try:
        subset = df[df["ID"].isin(ids_to_insert)].copy()
        # Asegurar tipos string para agrupación limpia
        subset["NIT"] = subset["NIT"].astype(str)
        subset["RAZON_SOCIAL"] = subset["RAZON_SOCIAL"].astype(str)

        def _d(col):
            return subset[col].apply(lambda x: parse_decimal_maybe_comma(str(x))).sum()

        grp = subset.groupby(["NIT", "RAZON_SOCIAL"], dropna=False)
        rows = []
        for (nit, razon), g in grp:
            cantidad = int(len(g))
            try:
                glosa = g["VALOR_GLOSA"].apply(lambda x: parse_decimal_maybe_comma(str(x))).sum()
            except Exception:
                glosa = Decimal("0")
            try:
                ips = g["VALOR_ACEPTADO_POR_IPS"].apply(lambda x: parse_decimal_maybe_comma(str(x))).sum()
            except Exception:
                ips = Decimal("0")
            try:
                eps = g["VALOR_ACEPTADO_POR_EPS"].apply(lambda x: parse_decimal_maybe_comma(str(x))).sum()
            except Exception:
                eps = Decimal("0")
            try:
                rat = g["VALOR_RATIFICADO_EPS"].apply(lambda x: parse_decimal_maybe_comma(str(x))).sum()
            except Exception:
                rat = Decimal("0")
            rows.append([
                str(nit or ""),
                str(razon or ""),
                cantidad,
                f"{glosa}",
                f"{ips}",
                f"{eps}",
                f"{rat}",
            ])

        out_path = os.path.join(out_dir, "dry_run_summary.csv")
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([
                "NIT",
                "RAZON_SOCIAL",
                "CANTIDAD_FACTURAS",
                "VALOR_GLOSA",
                "VALOR_ACEPTADO_POR_IPS",
                "VALOR_ACEPTADO_POR_EPS",
                "VALOR_RATIFICADO_EPS",
            ])
            for row in rows:
                w.writerow(row)
        return out_path
    except Exception as exc:
        error(f"No se pudo generar resumen de dry-run: {exc}")
        return ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", required=True, help="Ruta del CSV de entrada")
    ap.add_argument("--out-dir", required=True, help="Directorio de salida para reportes")
    ap.add_argument("--verbose", action="store_true", help="Muestra logs detallados en consola")
    ap.add_argument("--dry-run", action="store_true", help="Solo validar y generar resumen; no inserta en MySQL")
    ap.add_argument("--log-file", default="/data/out/runtime.log",
                    help="Ruta del log (por defecto /data/out/runtime.log si existe /data/out)")
    args = ap.parse_args()

    # preparar salida y log
    try:
        os.makedirs(args.out_dir, exist_ok=True)
    except Exception:
        pass
    log_file = args.log_file if os.path.isdir(os.path.dirname(args.log_file)) else None
    setup_logger(args.verbose, log_file)

    info("=== Inicio del proceso ===")
    info(f"CSV: {args.csv}")
    info(f"Salida: {args.out_dir}")
    if log_file:
        info(f"Log a archivo: {log_file}")

    # ==== Lectura robusta del CSV: auto-separador + BOM-safe ====
    try:
        debug("Intentando leer CSV con auto-detección (utf-8-sig, sep=None, engine=python)…")
        df = pd.read_csv(
            args.csv,
            dtype=str,
            keep_default_na=False,
            na_values=[],
            encoding="utf-8-sig",   # quita BOM si existe
            sep=None,               # auto-sniff del separador
            engine="python",
        )
    except Exception as e1:
        debug(f"Sniffer/utf-8-sig falló: {e1}. Intentando latin-1 y ';'…")
        try:
            df = pd.read_csv(
                args.csv,
                dtype=str,
                keep_default_na=False,
                na_values=[],
                encoding="latin-1",
                sep=";",
                engine="python",
            )
        except Exception as e2:
            error(f"No se pudo leer el CSV: {e2}")
            info("=== Fin del proceso con error de lectura CSV ===")
            sys.exit(EXIT_CSV_READ_ERROR)

    # Normalizar encabezados (remover espacios y BOM residual)
    headers_in_file = [ (h or "").strip().lstrip("\ufeff") for h in list(df.columns) ]
    df.columns = headers_in_file
    debug(f"Encabezados detectados (normalizados): {headers_in_file}")

    # Validaciones
    info("Validando estructura y reglas (CSV)…")
    errors = validate_csv(df, headers_in_file)

    if errors:
        out_err = os.path.join(args.out_dir, "log_errores.csv")
        write_errors_csv(errors, out_err)
        error(f"Se encontraron {len(errors)} errores. Ver: {out_err}")
        info("=== Fin del proceso con errores de CSV ===")
        sys.exit(EXIT_CSV_ERRORS)

    info("Validaciones CSV OK ✅")

    # Fase MySQL o DRY-RUN
    ids = df["ID"].tolist()

    if args.dry_run:
        info("DRY-RUN habilitado: se validará conexión MySQL y se generará resumen, sin insertar.")
        conn, err = get_mysql_conn_from_env()
        already = set()
        if conn is None:
            info(f"No se validará contra MySQL: {err}")
        else:
            try:
                with conn.cursor() as cur:
                    info("Verificando conciliaciones existentes (dry-run)…")
                    chunk_size = 1000
                    for i in range(0, len(ids), chunk_size):
                        chunk = ids[i:i+chunk_size]
                        placeholders = ",".join(["%s"] * len(chunk))
                        cur.execute(
                            f"SELECT auditory_final_report_id FROM conciliation_results WHERE auditory_final_report_id IN ({placeholders})",
                            chunk
                        )
                        already.update(row[0] for row in cur.fetchall())
            except Exception as ex:
                error(f"No fue posible consultar conciliaciones existentes en dry-run: {ex}")
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        ids_to_insert = [i for i in ids if i not in already]
        if already:
            info(f"IDs ya conciliados detectados en dry-run: {len(already)}; serán excluidos del resumen.")
        else:
            info("No se detectaron IDs ya conciliados en dry-run.")

        summary_path = _write_dry_run_summary(df, args.out_dir, ids_to_insert)
        if summary_path:
            info(f"Resumen de dry-run generado en: {summary_path}")
        info("=== Fin de DRY-RUN ===")
        sys.exit(EXIT_OK)

    info(f"Validando contra MySQL e insertando masivamente (IDs a procesar: {len(ids)})…")
    ok, msg, success_csv = mysql_phase_and_success_report(df, args.out_dir, ids)

    if ok:
        info(msg)
        if success_csv:
            info(f"Archivo de éxito generado: {success_csv}")
        info("=== Fin del proceso OK ===")
        sys.exit(EXIT_OK)
    else:
        error(f"Fallo en fase MySQL: {msg}")
        info("=== Fin del proceso con errores de MySQL ===")
        sys.exit(EXIT_MYSQL_ERROR)

if __name__ == "__main__":
    main()
