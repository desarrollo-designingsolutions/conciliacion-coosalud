#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cargar nuevas facturas: validación CSV, inserts en BD y generación de sábana Excel.

Uso desde app_streamlit.py (no es CLI standalone).
"""

import csv
import os
import uuid
import logging
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

try:
    import pymysql
except Exception:
    pymysql = None

# ------------------------ Constantes ------------------------

TZ = ZoneInfo("America/Bogota")

EXPECTED_HEADERS = [
    "nit",
    "razon_social",
    "factura",
    "modalidad",
    "fecha_prestacion",
    "fecha_factura",
    "fecha_radicacion",
    "valor_factura",
    "glosa",
    "radicacion",
]

COMPANY_ID = "9e5aec58-a962-4670-8188-b41c6d0149a3"
SERVICIO_ID = "31b734de-e29d-4e34-ad7c-83809ac1d32a"
VALID_RADICACION = {"nueva radicacion", "antigua"}

SABANA_OUTPUT_DIRNAME = "sabanas_nuevas_facturas"

# ------------------------ Logging helpers ------------------------

def info(msg: str): logging.info(msg)
def debug(msg: str): logging.debug(msg)
def error(msg: str): logging.error(msg)

# ------------------------ Modelos ------------------------

@dataclass
class NewInvoiceError:
    consecutivo: int
    fila: int
    columna: str
    dato: str
    error: str
    fecha: str


def now_bogota_str() -> str:
    return datetime.now(TZ).strftime("%d-%m-%Y %H:%M:%S")


def parse_decimal_maybe_comma(x: str) -> Decimal:
    if x is None:
        raise InvalidOperation("valor nulo")
    s = str(x).strip()
    if s == "":
        raise InvalidOperation("vacío")
    s = s.replace(" ", "").replace("$", "")
    if "," in s:
        s2 = s.replace(".", "").replace(",", ".")
        return Decimal(s2)
    return Decimal(s)


# ------------------------ Escritura de errores ------------------------

def write_errors_csv(errors: list[NewInvoiceError], out_path: str) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["consecutivo", "fila", "columna", "dato", "error", "fecha"])
        for e in errors:
            w.writerow([e.consecutivo, e.fila, e.columna, e.dato, e.error, e.fecha])


# ------------------------ Lectura del CSV ------------------------

def read_new_invoices_csv(file_path: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(
            file_path,
            dtype=str,
            keep_default_na=False,
            na_values=[],
            encoding="utf-8-sig",
            sep=";",
            engine="python",
        )
    except Exception:
        df = pd.read_csv(
            file_path,
            dtype=str,
            keep_default_na=False,
            na_values=[],
            encoding="latin-1",
            sep=";",
            engine="python",
        )
    headers = [(h or "").strip().lstrip("\ufeff") for h in list(df.columns)]
    df.columns = headers
    return df


# ------------------------ Validación CSV ------------------------

def validate_new_invoices_csv(
    df: pd.DataFrame,
    headers_in_file: list[str],
    conn=None,
) -> list[NewInvoiceError]:
    errors: list[NewInvoiceError] = []
    consecutivo = 1

    def add_error(fila: int, columna: str, dato: str, mensaje: str):
        nonlocal consecutivo
        errors.append(NewInvoiceError(
            consecutivo, fila, columna, dato, mensaje, now_bogota_str()
        ))
        consecutivo += 1

    # 1) Encabezados exactos y en orden
    if headers_in_file != EXPECTED_HEADERS:
        add_error(
            1, "HEADERS", "|".join(headers_in_file),
            f"Encabezados inválidos. Se esperan: {';'.join(EXPECTED_HEADERS)}"
        )
        return errors

    # 2) Campos obligatorios (no vacíos)
    for idx, row in df.iterrows():
        fila = idx + 2
        for col in EXPECTED_HEADERS:
            val = str(row.get(col, "")).strip()
            if val == "":
                add_error(fila, col, "", f"El campo '{col}' es obligatorio y está vacío.")

    if errors:
        return errors

    # 3) NIT: un solo valor único, numérico
    nits_unicos = df["nit"].astype(str).str.strip().unique()
    if len(nits_unicos) != 1:
        add_error(
            0, "nit", "|".join(nits_unicos),
            f"El archivo debe contener un único NIT. Se encontraron {len(nits_unicos)}: {', '.join(nits_unicos)}"
        )
        return errors

    nit_val = nits_unicos[0]
    if not nit_val.isdigit():
        add_error(0, "nit", nit_val, "El NIT debe ser numérico.")
        return errors

    # NIT debe existir en tabla thirds
    if conn is not None:
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT id, name FROM thirds WHERE id = %s", (nit_val,))
                third_row = cur.fetchone()
                if third_row is None:
                    add_error(0, "nit", nit_val, f"El NIT {nit_val} no existe en la tabla thirds.")
                    return errors
        except Exception as exc:
            add_error(0, "nit", nit_val, f"Error consultando NIT en thirds: {exc}")
            return errors

    # 4) Facturas: sin duplicados en el archivo
    facturas = df["factura"].astype(str).str.strip()
    duplicadas = facturas[facturas.duplicated(keep=False)]
    if not duplicadas.empty:
        seen = set()
        for idx, val in duplicadas.items():
            if val in seen:
                continue
            seen.add(val)
            filas_dup = [str(i + 2) for i, v in duplicadas.items() if v == val]
            add_error(
                int(idx) + 2, "factura", val,
                f"Factura duplicada en el archivo. Aparece en filas: {', '.join(filas_dup)}"
            )

    # Facturas no deben existir en BD
    if conn is not None:
        facturas_list = facturas.unique().tolist()
        if facturas_list:
            try:
                with conn.cursor() as cur:
                    placeholders = ",".join(["%s"] * len(facturas_list))
                    cur.execute(
                        f"SELECT DISTINCT ia.invoice_number FROM invoice_audits ia "
                        f"INNER JOIN auditory_final_reports afr ON afr.factura_id = ia.id "
                        f"INNER JOIN estados_auditory_final_reports eafr ON eafr.ID = afr.id "
                        f"WHERE ia.third_id = %s AND eafr.ESTADO = 'contabilizada' "
                        f"AND ia.invoice_number IN ({placeholders})",
                        [nit_val] + facturas_list,
                    )
                    existentes = {row[0] for row in cur.fetchall()}
                    for idx, row in df.iterrows():
                        fac = str(row["factura"]).strip()
                        if fac in existentes:
                            add_error(
                                idx + 2, "factura", fac,
                                f"La factura '{fac}' ya existe en la base de datos para el NIT {nit_val}."
                            )
            except Exception as exc:
                add_error(0, "factura", "", f"Error consultando facturas existentes: {exc}")

    # 5) Fechas: formato yyyy-mm-dd
    date_cols = ["fecha_prestacion", "fecha_factura", "fecha_radicacion"]
    for idx, row in df.iterrows():
        fila = idx + 2
        for col in date_cols:
            val = str(row.get(col, "")).strip()
            if val == "":
                continue  # ya validado en obligatorios
            try:
                datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                add_error(fila, col, val, f"Formato de fecha inválido. Se espera yyyy-mm-dd.")

    # 6) valor_factura: numérico
    for idx, row in df.iterrows():
        fila = idx + 2
        val = str(row.get("valor_factura", "")).strip()
        if val == "":
            continue
        try:
            parse_decimal_maybe_comma(val)
        except (InvalidOperation, Exception):
            add_error(fila, "valor_factura", val, "valor_factura debe ser numérico.")

    # 7) glosa: numérico, > 0, <= valor_factura
    for idx, row in df.iterrows():
        fila = idx + 2
        glosa_str = str(row.get("glosa", "")).strip()
        vf_str = str(row.get("valor_factura", "")).strip()
        if glosa_str == "":
            continue

        try:
            glosa_dec = parse_decimal_maybe_comma(glosa_str)
        except (InvalidOperation, Exception):
            add_error(fila, "glosa", glosa_str, "glosa debe ser numérico.")
            continue

        if glosa_dec <= 0:
            add_error(fila, "glosa", glosa_str, "glosa debe ser mayor que cero.")

        try:
            vf_dec = parse_decimal_maybe_comma(vf_str)
            if glosa_dec > vf_dec:
                add_error(
                    fila, "glosa", glosa_str,
                    f"glosa ({glosa_str}) no puede ser mayor que valor_factura ({vf_str})."
                )
        except (InvalidOperation, Exception):
            pass  # error de valor_factura ya reportado arriba

    # 8) radicacion: solo valores permitidos
    for idx, row in df.iterrows():
        fila = idx + 2
        val = str(row.get("radicacion", "")).strip()
        if val == "":
            continue
        if val not in VALID_RADICACION:
            add_error(
                fila, "radicacion", val,
                f"radicacion inválida. Solo se acepta: {', '.join(sorted(VALID_RADICACION))}"
            )

    return errors


# ------------------------ Conexión BD ------------------------

def get_mysql_conn():
    host = os.getenv("MYSQL_HOST")
    port = int(os.getenv("MYSQL_PORT", "3306"))
    db = os.getenv("MYSQL_DB")
    user = os.getenv("MYSQL_USER")
    pw = os.getenv("MYSQL_PASSWORD")
    if not all([host, db, user, pw]):
        return None, "Variables de entorno MySQL incompletas (MYSQL_HOST, MYSQL_DB, MYSQL_USER, MYSQL_PASSWORD)"
    if pymysql is None:
        return None, "pymysql no instalado"
    try:
        conn = pymysql.connect(
            host=host, port=port, user=user, password=pw,
            database=db, autocommit=False, charset="utf8mb4",
            cursorclass=pymysql.cursors.Cursor,
        )
        with conn.cursor() as cur:
            cur.execute("SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci")
            cur.execute("SET collation_connection = 'utf8mb4_unicode_ci'")
        conn.commit()
        return conn, None
    except Exception as ex:
        return None, f"No fue posible conectar a MySQL: {ex}"


# ------------------------ Inserts en BD ------------------------

def build_origin(nit: str) -> str:
    fecha = datetime.now(TZ).strftime("%Y%m%d")
    return f"{nit}_nuevas_{fecha}"


def insert_new_invoices(df: pd.DataFrame, conn, nit: str, origin: str) -> tuple[bool, str, int]:
    """
    Inserta en invoice_audits, auditory_final_reports y estados_auditory_final_reports.
    Retorna (ok, mensaje, cantidad_insertada).
    """
    fecha_proceso = datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")

    try:
        with conn.cursor() as cur:
            # --- 1) Tabla temporal con datos del CSV ---
            cur.execute("""
                CREATE TEMPORARY TABLE IF NOT EXISTS tmp_new_invoices (
                    nit VARCHAR(50) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL,
                    razon_social VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL,
                    factura VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL,
                    modalidad VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NULL,
                    fecha_prestacion DATE NULL,
                    fecha_factura DATE NULL,
                    fecha_radicacion DATE NULL,
                    valor_factura DECIMAL(18,2) NOT NULL,
                    glosa DECIMAL(18,2) NOT NULL,
                    radicacion VARCHAR(50) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NULL,
                    PRIMARY KEY (factura)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            cur.execute("TRUNCATE TABLE tmp_new_invoices")

            # Cargar datos
            insert_tmp_sql = """
                INSERT INTO tmp_new_invoices
                (nit, razon_social, factura, modalidad, fecha_prestacion, fecha_factura, fecha_radicacion, valor_factura, glosa, radicacion)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            tmp_rows = []
            for _, row in df.iterrows():
                vf = parse_decimal_maybe_comma(str(row["valor_factura"]))
                gl = parse_decimal_maybe_comma(str(row["glosa"]))
                tmp_rows.append((
                    str(row["nit"]).strip(),
                    str(row["razon_social"]).strip(),
                    str(row["factura"]).strip(),
                    str(row["modalidad"]).strip(),
                    str(row["fecha_prestacion"]).strip(),
                    str(row["fecha_factura"]).strip(),
                    str(row["fecha_radicacion"]).strip(),
                    str(vf),
                    str(gl),
                    str(row["radicacion"]).strip(),
                ))

            batch_size = 5000
            for i in range(0, len(tmp_rows), batch_size):
                chunk = tmp_rows[i:i + batch_size]
                cur.executemany(insert_tmp_sql, chunk)
            conn.commit()
            info(f"Tabla temporal cargada: {len(tmp_rows)} filas.")

            # --- 2) INSERT en invoice_audits ---
            cur.execute(f"""
                INSERT INTO invoice_audits
                (id, company_id, third_id, filing_invoice_id, invoice_number,
                 total_value, origin, expedition_date, date_entry, date_departure,
                 modality, regimen, coverage, contract_number,
                 created_at, updated_at, deleted_at, status_radication, radication)
                SELECT
                    UUID(),
                    %s,
                    nf.nit,
                    NULL,
                    nf.factura,
                    nf.valor_factura,
                    %s,
                    nf.fecha_factura,
                    nf.fecha_factura,
                    nf.fecha_factura,
                    nf.modalidad,
                    'SUBSIDIADO',
                    'POS',
                    NULL,
                    NOW(),
                    NOW(),
                    NULL,
                    'devolución aplistaf',
                    nf.radicacion
                FROM tmp_new_invoices nf
            """, (COMPANY_ID, origin))
            inserted_ia = cur.rowcount
            conn.commit()
            info(f"invoice_audits insertados: {inserted_ia}")

            # --- 3) INSERT en auditory_final_reports ---
            cur.execute(f"""
                INSERT INTO auditory_final_reports
                (id, factura_id, servicio_id, origin, nit, razon_social,
                 numero_factura, fecha_inicio, fecha_fin, modalidad, regimen, cobertura, contrato,
                 tipo_documento, numero_documento, primer_nombre, segundo_nombre,
                 primer_apellido, segundo_apellido, genero,
                 codigo_servicio, descripcion_servicio, cantidad_servicio,
                 valor_unitario_servicio, valor_total_servicio,
                 codigos_glosa, observaciones_glosas, valor_glosa, valor_aprobado,
                 created_at, updated_at, deleted_at)
                SELECT
                    UUID(),
                    ia.id,
                    %s,
                    ia.origin,
                    t.id,
                    t.name,
                    ia.invoice_number,
                    ia.expedition_date,
                    ia.expedition_date,
                    ia.modality,
                    ia.regimen,
                    ia.coverage,
                    ia.contract_number,
                    'CC',
                    '000000',
                    'NA',
                    'NA',
                    'NA',
                    'NA',
                    'M',
                    '000000',
                    'SERVICIOS FACTURADOS',
                    1,
                    ia.total_value,
                    ia.total_value,
                    '223',
                    'la factura presenta diferencias con los valores que fueron pactados',
                    nf.glosa,
                    (ia.total_value - nf.glosa),
                    NOW(),
                    NOW(),
                    NULL
                FROM invoice_audits ia
                INNER JOIN thirds t
                    ON t.id COLLATE utf8mb4_unicode_ci = ia.third_id COLLATE utf8mb4_unicode_ci
                INNER JOIN tmp_new_invoices nf
                    ON nf.factura COLLATE utf8mb4_unicode_ci = ia.invoice_number COLLATE utf8mb4_unicode_ci
                    AND nf.nit COLLATE utf8mb4_unicode_ci = t.id COLLATE utf8mb4_unicode_ci
                WHERE t.id = %s
                AND ia.created_at > %s
                AND ia.origin = %s
            """, (SERVICIO_ID, nit, fecha_proceso, origin))
            inserted_afr = cur.rowcount
            conn.commit()
            info(f"auditory_final_reports insertados: {inserted_afr}")

            # --- 4) INSERT en estados_auditory_final_reports ---
            cur.execute("""
                INSERT INTO estados_auditory_final_reports (ID, FACTURA_ID, ESTADO)
                SELECT
                    afr.id,
                    afr.factura_id,
                    'contabilizada'
                FROM auditory_final_reports afr
                WHERE afr.nit = %s
                AND afr.created_at > %s
                AND afr.origin = %s
            """, (nit, fecha_proceso, origin))
            inserted_eafr = cur.rowcount
            conn.commit()
            info(f"estados_auditory_final_reports insertados: {inserted_eafr}")

        return True, f"Proceso exitoso. invoice_audits: {inserted_ia}, auditory_final_reports: {inserted_afr}, estados: {inserted_eafr}", inserted_ia

    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        error(f"Error durante inserts: {exc}")
        return False, f"Error durante la inserción en BD: {exc}", 0


# ------------------------ Generación de sábana Excel ------------------------

def generate_sabana_excel(conn, nit: str, origin: str, fecha_proceso: str, out_dir: str) -> tuple[bool, str]:
    """
    Genera sábana Excel con las facturas recién insertadas.
    Retorna (ok, file_path).
    """
    sabana_dir = Path(out_dir) / SABANA_OUTPUT_DIRNAME
    try:
        sabana_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        return False, f"No se pudo crear directorio: {exc}"

    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    aud.id,
                    aud.factura_id,
                    aud.servicio_id,
                    aud.origin,
                    aud.nit,
                    aud.razon_social,
                    aud.numero_factura,
                    aud.fecha_inicio,
                    aud.fecha_fin,
                    aud.modalidad,
                    aud.regimen,
                    aud.cobertura,
                    aud.contrato,
                    aud.tipo_documento,
                    aud.numero_documento,
                    aud.primer_nombre,
                    aud.segundo_nombre,
                    aud.primer_apellido,
                    aud.segundo_apellido,
                    aud.genero,
                    aud.codigo_servicio,
                    aud.descripcion_servicio,
                    aud.cantidad_servicio,
                    aud.valor_unitario_servicio,
                    aud.valor_total_servicio,
                    aud.codigos_glosa,
                    REGEXP_REPLACE(aud.observaciones_glosas, '[\\r\\n\\t]+', ' ') AS observaciones_glosas,
                    aud.valor_glosa,
                    aud.valor_aprobado,
                    NULL AS estado_respuesta,
                    NULL AS numero_de_autorizacion,
                    NULL AS respuesta_de_ips,
                    NULL AS valor_aceptado_por_ips,
                    NULL AS valor_aceptado_por_eps,
                    NULL AS valor_ratificado_eps,
                    NULL AS observaciones
                FROM auditory_final_reports aud
                WHERE aud.nit = %s
                AND aud.origin = %s
                AND aud.created_at > %s
                ORDER BY aud.id
            """, (nit, origin, fecha_proceso))
            rows = cur.fetchall()
            columns = [desc[0].upper() for desc in cur.description]

        if not rows:
            return False, "No se encontraron registros para generar la sábana."

        wb = Workbook()
        ws = wb.active
        ws.title = "Sabana"

        # Encabezados
        for col_idx, header in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Datos
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell_val = value
                if isinstance(value, Decimal):
                    cell_val = float(value)
                ws.cell(row=row_idx, column=col_idx, value=cell_val)

        timestamp = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")
        file_name = f"{nit}_sabana_{timestamp}.xlsx"
        file_path = str(sabana_dir / file_name)
        wb.save(file_path)
        wb.close()

        # Subir a S3 si está habilitado
        from s3_storage import s3_enabled, upload_file, build_s3_key
        stored_path = file_path
        if s3_enabled():
            s3_key = build_s3_key(f"sabanas_nuevas_facturas/{file_name}")
            s3_ok, s3_result = upload_file(file_path, s3_key)
            if s3_ok:
                stored_path = s3_key
            else:
                error(f"No se pudo subir sábana a S3: {s3_result}")

        info(f"Sábana generada: {stored_path}")
        return True, stored_path

    except Exception as exc:
        error(f"Error generando sábana: {exc}")
        return False, f"Error generando sábana: {exc}"


# ------------------------ Historial de importaciones ------------------------

CREATE_IMPORTS_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS new_invoice_imports (
  id CHAR(36) NOT NULL,
  nit VARCHAR(50) NOT NULL,
  razon_social VARCHAR(255) NOT NULL,
  origin VARCHAR(255) NOT NULL,
  total_facturas INT NOT NULL DEFAULT 0,
  valor_factura_total DECIMAL(18,2) NOT NULL DEFAULT 0,
  glosa_total DECIMAL(18,2) NOT NULL DEFAULT 0,
  file_name VARCHAR(255) NULL,
  file_path VARCHAR(500) NULL,
  usuario VARCHAR(255) NOT NULL,
  created_at DATETIME NOT NULL,
  PRIMARY KEY (id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
"""


def ensure_imports_table(conn):
    try:
        with conn.cursor() as cur:
            cur.execute(CREATE_IMPORTS_TABLE_SQL)
        conn.commit()
    except Exception as exc:
        error(f"Error creando tabla new_invoice_imports: {exc}")


def register_import(conn, nit: str, razon_social: str, origin: str,
                    total_facturas: int, valor_factura_total: Decimal,
                    glosa_total: Decimal, file_name: str, file_path: str,
                    usuario: str) -> str | None:
    import_id = str(uuid.uuid4())
    try:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO new_invoice_imports
                (id, nit, razon_social, origin, total_facturas, valor_factura_total,
                 glosa_total, file_name, file_path, usuario, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """, (
                import_id, nit, razon_social, origin, total_facturas,
                str(valor_factura_total), str(glosa_total),
                file_name, file_path, usuario,
            ))
        conn.commit()
        return import_id
    except Exception as exc:
        error(f"Error registrando importación: {exc}")
        return None
